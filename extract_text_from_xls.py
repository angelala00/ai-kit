from openpyxl import load_workbook

wb = load_workbook("/Users/chijiang/Downloads/Book1.xlsx")
sheet = wb.active
text = ""
for row in sheet.iter_rows(values_only=True):
    text += " ".join(str(cell) for cell in row) + "\n"
print(text)

